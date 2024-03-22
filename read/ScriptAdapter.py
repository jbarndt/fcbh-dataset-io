# This program will read Excel data and load the audio_scripts table

import re
import os
import csv
import sys
from openpyxl import load_workbook
from DBAdapter import *
from WordParser import *
from Booknames import *


class ScriptAdapter:

	def __init__(self, db):
		self.db = db
		self.bookNames = Booknames()


	def findFile(self, bibleId):
		directory = os.path.join(os.environ["FCBH_DATASET_FILES"], bibleId)
		for file in os.listdir(directory):
			if file.endswith(".xlsx"):
				return os.path.join(directory, file)
		print("Could not find .xlsx file in", directory)
		sys.exit(1)


	def load(self, filename):
		print("reading", filename)
		#tmpFile = open("temp.txt", "w")
		workbook = load_workbook(filename=filename)
		sheet_names = workbook.sheetnames
		wb = workbook[sheet_names[0]]
		for row in wb.iter_rows(min_row=2, max_row=wb.max_row, min_col=1, max_col=wb.max_column):
			#print("max col", wb.max_column)
			book_id = row[1].value
			if not self.bookNames.isBook(book_id):
				correctBook = {'TTS': 'TIT', 'JMS': 'JAS'}
				book_id = correctBook.get(book_id)
				if book_id == None:
					print("Found book_id", row[1].value)
					sys.exit(1)
			chapter_num = row[2].value
			audio_file = "xxxxxxxDA_" + book_id + "_" + str(chapter_num) + ".mp3"
			verse_str = row[3].value
			if verse_str == "<<":
				verse_str = ""
			in_verse_num = row[3].value
			if in_verse_num == "<<":
				#in_verse_num = None
				in_verse_num = 0
			usfm_style = None
			person = row[4].value
			#actor = row[5].value
			actor = None
			script_num = str(row[5].value)
			script_text = row[8].value.replace('_x000D_','') # remove excel CR
			if not script_num[-1] == 'r':
				self.db.addScript(book_id, chapter_num, audio_file, script_num, usfm_style, 
							person, actor, in_verse_num, verse_str, script_text)
		self.db.insertScripts()
		workbook.close()


if __name__ == "__main__":
	if len(sys.argv) < 2:
		print("Usage: python3 ScriptAdapter.py  bibleId")
		sys.exit(1)
	bibleId = sys.argv[1]
	database = bibleId + "_SCRIPT.db"
	DBAdapter.destroyDatabase(database)
	db = DBAdapter(database)
	script = ScriptAdapter(db)
	filename = script.findFile(bibleId)
	script.load(filename)
	word = WordParser(db)
	word.parse()
	db.close()

