# This program will read Excel data and load the audio_scripts table

import re
import os
import csv
from openpyxl import load_workbook
from DBAdapter import *
from FileAdapter import *



class ExcelAdapter:

	def __init__(self, db):
		self.db = db


	def load(self, filename):
		workbook = load_workbook(filename=filename)
		sheet_names = workbook.sheetnames
		wb = workbook[sheet_names[0]]
		script_num = 0
		for row in wb.iter_rows(min_row=3, max_row=wb.max_row, min_col=1, max_col=wb.max_column):
			#print("max col", wb.max_column)
			book_id = row[1].value
			chapter_num = row[2].value
			audio_file = "xxxxxxxDA_" + book_id + "_" + str(chapter_num) + ".mp3"
			in_verse_num = row[3].value
			if in_verse_num == "<<":
				in_verse_num = None
			#if not isinstance(chapter_num, int):
			#	print("chapter", type(chapter_num), chapter_num)
			#if not isinstance(in_verse_num, int):
			#	print("verse", type(in_verse_num), in_verse_num)
			usfm_style = None
			person = row[4].value
			actor = row[5].value
			script_text = row[8].value
			#print("6", row[6].value)
			#print("7", row[7].value)
			#print("8", row[8].value)
			script_num += 1
			self.db.addScript(book_id, chapter_num, audio_file, script_num, usfm_style, 
							person, actor, in_verse_num, script_text)
		self.db.insertScripts()



if __name__ == "__main__":
	database = "ZAK_MWRIGHT"
	DBAdapter.destroyDatabase(database)
	db = DBAdapter(database)
	exc = ExcelAdapter(db)
	filename = os.path.join(os.environ["HOME"], "FCBH2024", "Mark_Wright_Feb_2024", "Context ZAK.xlsx")
	print("Read:", filename)
	exc.load(filename)
	file = FileAdapter(db)
	file.loadWords()
