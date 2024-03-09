# This program will read Excel data and load the audio_scripts table

import re
import os
import csv
import sys
from openpyxl import load_workbook
from DBAdapter import *
from WordParser import *
from Booknames import *


class ScriptAdapter:

	def __init__(self, db):
		self.db = db
		self.bookNames = Booknames()


	def findFile(self, bibleId):
		directory = os.path.join(os.environ["FCBH_DATASET_FILES"], bibleId)
		for file in os.listdir(directory):
			if file.endswith(".xlsx"):
				return os.path.join(directory, file)
		print("Could not find .xlsx file in", directory)
		sys.exit(1)


	def load(self, filename):
		print("reading", filename)
		workbook = load_workbook(filename=filename)
		sheet_names = workbook.sheetnames
		wb = workbook[sheet_names[0]]
		for row in wb.iter_rows(min_row=2, max_row=wb.max_row, min_col=1, max_col=wb.max_column):
			#print("max col", wb.max_column)
			book_id = row[1].value
			if not self.bookNames.isBook(book_id):
				correctBook = {'TTS': 'TIT', 'JMS': 'JAS'}
				book_id = correctBook.get(book_id)
				if book_id == None:
					print("Found book_id", row[1].value)
					sys.exit(1)
			chapter_num = row[2].value
			audio_file = "xxxxxxxDA_" + book_id + "_" + str(chapter_num) + ".mp3"
			in_verse_num = row[3].value
			if in_verse_num == "<<":
				in_verse_num = None
			usfm_style = None
			person = row[4].value
			#actor = row[5].value
			script_num = str(row[5].value)
			script_text = row[8].value
			actor = None
			if script_num[-1].isdigit():
				self.db.addScript(book_id, chapter_num, audio_file, script_num, usfm_style, 
							person, actor, in_verse_num, script_text)
		self.db.insertScripts()


if __name__ == "__main__":
	if len(sys.argv) < 2:
		print("Usage: python3 ExcelAdapter.py  bibleId")
		sys.exit(1)
	bibleId = sys.argv[1]
	database = bibleId + "_SCRIPT.db"
	DBAdapter.destroyDatabase(database)
	db = DBAdapter(database)
	script = ScriptAdapter(db)
	filename = script.findFile(bibleId)
	script.load(filename)
	word = WordParser(db)
	word.parse()
	db.close()

